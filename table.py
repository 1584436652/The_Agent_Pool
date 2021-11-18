from openpyxl import load_workbook
from openpyxl import Workbook


def load_table():
    wb = load_workbook('re.xlsx')
    ws = wb["整理后"]
    # ws = wb.active
    rows = []
    sku_dict = dict()
    for row in ws.iter_rows():
        rows.append(row)
    for x in range(1, len(rows)):
        sku_name = str(rows[x][0].value)
        billing_weight = float(rows[x][3].value)
        if sku_name not in sku_dict:
            sku_dict[sku_name] = [billing_weight]
        else:
            sku_dict[sku_name].append(billing_weight)
    return sku_dict


def billing_weight_sort(items):
    data = []
    for k, v in items.items():
        # 计费重排序
        so = sorted(v)
        # 最小计费重
        min_v = float(so[0])
        # 最大计费重
        max_v = float(so[-1])
        # 中位数
        size = len(so)
        if size % 2 == 0:
            median = float(format((so[size // 2] + so[size // 2 - 1]) / 2, '.3f'))
        else:
            median = float(so[(size - 1) // 2])
        if max_v - min_v >= median:
            abnormal = [k, so, min_v, max_v, median, "abnormal"]
            data.append(abnormal)
        else:
            abnormal = [k, so, min_v, max_v, median, None]
            data.append(abnormal)
    return data


def save(save_data):
    print(save_data)
    wb = Workbook()
    ws = wb.active
    ws.append(["SKU", "SORT", "MIN", "MAX", "MEDIAN", "BOOL"])
    print(f'总条数：{len(save_data)}')
    number = 2
    for save_value in save_data:
        ws[f'A{number}'] = save_value[0]
        ws[f'B{number}'] = str(save_value[1])
        ws[f'C{number}'] = save_value[2]
        ws[f'D{number}'] = save_value[3]
        ws[f'E{number}'] = save_value[4]
        ws[f'F{number}'] = save_value[5]
        wb.save('demo.xlsx')
        number += 1
    print("已保存")


if __name__ == '__main__':
    it = load_table()
    sot = billing_weight_sort(it)
    save(sot)



